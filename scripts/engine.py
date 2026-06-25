"""
Felles parser + poengmotor for Piggy's tippekonkurranser.
Brukes på tvers av alle turneringer (VM 2026, EM 2028, ...).

Leser et utfylt "Excely"-ark (arket "Resultater") og trekker ut alle
prediksjoner. Regner poeng mot en fasit av samme form. Ren stdlib +
openpyxl — ingen tunge avhengigheter.
"""
from __future__ import annotations
import json
from pathlib import Path
from openpyxl import load_workbook

# Faste celleadresser i "Resultater"-arket (1-indeksert: rad, kolonne).
# Disse er stabile i Excely-malen og deles av alle turneringer som bruker den.
LAYOUT = {
    "match_rows": (4, 75), "col_kamp": 2, "col_home": 3, "col_away": 4, "col_result": 5,
    "gw_rows": (78, 89), "col_gw_label": 4, "col_gw_team": 5,
    "r16_rows": (92, 123), "r8_rows": (126, 141), "kvart_rows": (144, 151),
    "semi_rows": (154, 157), "bronse_rows": (160, 161), "bronse_vinner_row": 162,
    "finale_rows": (165, 166), "vm_vinner_row": 167, "col_ko_team": 5,
    "bonus_toppscorer": (26, 11), "bonus_assist": (27, 11),
    "bonus_maal": (28, 11), "bonus_kort": (29, 11),
}


def _norm(s) -> str:
    return str(s).strip().lower() if s is not None else ""


def _col_block(ws, rows, col):
    out = []
    for r in range(rows[0], rows[1] + 1):
        v = ws.cell(r, col).value
        if v is not None and str(v).strip():
            out.append(str(v).strip())
    return out


def parse_sheet(path: str | Path) -> dict:
    """Trekk ut alle prediksjoner fra ett utfylt ark."""
    wb = load_workbook(path, data_only=True)
    if "Resultater" not in wb.sheetnames:
        raise ValueError(f'{Path(path).name}: fant ikke arket "Resultater"')
    ws = wb["Resultater"]
    L = LAYOUT

    matches = []
    for r in range(L["match_rows"][0], L["match_rows"][1] + 1):
        n = ws.cell(r, L["col_kamp"]).value
        if n is None:
            continue
        matches.append({
            "n": int(n),
            "home": str(ws.cell(r, L["col_home"]).value or "").strip(),
            "away": str(ws.cell(r, L["col_away"]).value or "").strip(),
            "pick": str(ws.cell(r, L["col_result"]).value or "").strip().upper(),
        })

    gw = {}
    for r in range(L["gw_rows"][0], L["gw_rows"][1] + 1):
        label = ws.cell(r, L["col_gw_label"]).value
        team = ws.cell(r, L["col_gw_team"]).value
        if label and team:
            gw[str(label).strip()] = str(team).strip()

    def cell(rc):
        return ws.cell(rc[0], rc[1]).value

    maal = cell(L["bonus_maal"])
    kort = cell(L["bonus_kort"])
    return {
        "matches": matches,
        "group_winners": gw,
        "r16": _col_block(ws, L["r16_rows"], L["col_ko_team"]),
        "r8": _col_block(ws, L["r8_rows"], L["col_ko_team"]),
        "kvart": _col_block(ws, L["kvart_rows"], L["col_ko_team"]),
        "semi": _col_block(ws, L["semi_rows"], L["col_ko_team"]),
        "bronse": _col_block(ws, L["bronse_rows"], L["col_ko_team"]),
        "bronse_vinner": str(cell((L["bronse_vinner_row"], L["col_ko_team"])) or "").strip(),
        "finale": _col_block(ws, L["finale_rows"], L["col_ko_team"]),
        "vm_vinner": str(cell((L["vm_vinner_row"], L["col_ko_team"])) or "").strip(),
        "toppscorer": str(cell(L["bonus_toppscorer"]) or "").strip(),
        "assist": str(cell(L["bonus_assist"]) or "").strip(),
        "antall_maal": int(maal) if maal not in (None, "") else None,
        "antall_kort": int(kort) if kort not in (None, "") else None,
    }


def parse_master(path: str | Path) -> list[dict]:
    """Parse MASTER-arket og returner en liste med prediksjoner per deltaker."""
    wb = load_workbook(path, data_only=True)
    ws = wb["MASTER"]

    # Deltakernavn i rad 3, kolonne 5-15
    deltakere = []
    for c in range(5, 16):
        navn = ws.cell(3, c).value
        if navn:
            deltakere.append({"col": c, "navn": str(navn).strip()})

    results = []
    for d in deltakere:
        col = d["col"]

        # Kamper rad 4-75
        matches = []
        for r in range(4, 76):
            n = ws.cell(r, 2).value
            home = str(ws.cell(r, 3).value or "").strip()
            away = str(ws.cell(r, 4).value or "").strip()
            pick = str(ws.cell(r, col).value or "").strip().upper()
            if n is not None and home:
                matches.append({"n": int(n), "home": home, "away": away, "pick": pick})

        # Gruppevinnere rad 77-88
        gw = {}
        for r in range(77, 89):
            label = ws.cell(r, 3).value
            team = ws.cell(r, col).value
            if label and team:
                gw[str(label).strip()] = str(team).strip()

        # Sluttspill - lese lag fra deltaker-kolonne
        def col_block(start, end):
            out = []
            for r in range(start, end + 1):
                v = ws.cell(r, col).value
                if v is not None and str(v).strip():
                    out.append(str(v).strip())
            return out

        r16 = col_block(91, 122)
        r8 = col_block(125, 140)
        kvart = col_block(143, 150)
        semi = col_block(153, 156)
        bronse = col_block(153, 156)  # same as semi for bronsefinale-lag

        # Bronsefinale: rad 151-152 for lag, rad 153 for vinner? Let me check layout
        bronse_lag = col_block(153, 154)
        bronse_vinner = str(ws.cell(155, col).value or "").strip()
        finale_lag = col_block(158, 159)
        vm_vinner = str(ws.cell(160, col).value or "").strip()

        # Bonus
        toppscorer = str(ws.cell(162, col).value or "").strip()
        assist = str(ws.cell(163, col).value or "").strip()
        maal = ws.cell(164, col).value
        kort = ws.cell(165, col).value

        results.append({
            "navn": d["navn"],
            "pred": {
                "matches": matches,
                "group_winners": gw,
                "r16": r16,
                "r8": r8,
                "kvart": kvart,
                "semi": semi,
                "bronse": bronse_lag,
                "bronse_vinner": bronse_vinner,
                "finale": finale_lag,
                "vm_vinner": vm_vinner,
                "toppscorer": toppscorer,
                "assist": assist,
                "antall_maal": int(maal) if maal not in (None, "") else None,
                "antall_kort": int(kort) if kort not in (None, "") else None,
            }
        })

    return results


def _count_set(preds, fact_set):
    hits, used = 0, set()
    for p in preds:
        k = _norm(p)
        if k in fact_set and k not in used:
            hits += 1
            used.add(k)
    return hits


def score(pred: dict, fact: dict, rules: dict) -> dict:
    """Regn poeng for én prediksjon mot fasit. `rules` = poengverdier."""
    lines, total = [], 0

    def add(key, label, pts, detail, kamper=None, grupper=None):
        nonlocal total
        line = {"key": key, "label": label, "pts": pts, "detail": detail}
        if kamper:
            line["kamper"] = kamper
        if grupper:
            line["grupper"] = grupper
        lines.append(line)
        total += pts

    # Gruppespill (H/U/B). Match by home+away team names.
    if fact.get("matches"):
        fact_by_teams = {(_norm(m["home"]), _norm(m["away"])): _norm(m["result"]) for m in fact["matches"]}
        treff = []
        for m in pred["matches"]:
            res = fact_by_teams.get((_norm(m["home"]), _norm(m["away"])))
            if res and _norm(m["pick"]) == res:
                treff.append({"home": m["home"], "away": m["away"], "res": res.upper()})
        add("hub", f"Gruppespill ({len(treff)} riktige)", len(treff) * rules["kamp"],
            f'{len(treff)} x {rules["kamp"]}p', kamper=treff)

    # Gruppevinnere: gis per FERDIGSPILT gruppe (en ferdig gruppe er låst, så
    # poenget er stabilt). Uferdige grupper vises ikke og teller ikke.
    grupper_ferdig = fact.get("grupper_ferdig")
    if not grupper_ferdig:
        # Utled fra kampene hvis fetch ikke har lagt feltet inn ennå.
        g_tot, g_fin = {}, {}
        for k in fact.get("kamper", []):
            if k.get("stage") != "GROUP_STAGE" or not k.get("group"):
                continue
            lab = "Gruppe " + k["group"].split("_")[-1]
            g_tot[lab] = g_tot.get(lab, 0) + 1
            if k.get("status") == "FINISHED" and k.get("home_score") is not None:
                g_fin[lab] = g_fin.get(lab, 0) + 1
        grupper_ferdig = {lab: g_fin.get(lab, 0) == g_tot[lab] for lab in g_tot}
        if not grupper_ferdig and fact.get("gruppespill_ferdig"):
            grupper_ferdig = {g: True for g in (fact.get("group_winners") or {})}
    if fact.get("group_winners") and any(grupper_ferdig.values()):
        gr_detalj, correct = [], 0
        for g, fasit_v in fact["group_winners"].items():
            if not grupper_ferdig.get(g):
                continue
            tipp = pred["group_winners"].get(g, "")
            hit = bool(fasit_v) and _norm(fasit_v) == _norm(tipp)
            if hit:
                correct += 1
            gr_detalj.append({"gruppe": g, "fasit": fasit_v, "tipp": tipp or "-", "hit": hit})
        add("gruppevinner", f"Gruppevinnere ({correct} riktige)",
            correct * rules["gruppevinner"], f'{correct} x {rules["gruppevinner"]}p', grupper=gr_detalj)

    rounds = [
        ("16-dels finale", "r16", "r16"), ("8-dels finale", "r8", "r8"),
        ("Kvartfinale", "kvart", "kvart"), ("Semifinale", "semi", "semi"),
        ("Bronsefinale (lag)", "bronse", "bronse_lag"), ("Finale (lag)", "finale", "finale_lag"),
    ]
    for label, key, rule_key in rounds:
        if not fact.get(key):
            continue
        hits = _count_set(pred.get(key, []), {_norm(x) for x in fact[key]})
        add(rule_key, f"{label} ({hits} riktige)", hits * rules[rule_key], f'{hits} x {rules[rule_key]}p')

    if fact.get("bronse_vinner"):
        ok = _norm(pred.get("bronse_vinner")) == _norm(fact["bronse_vinner"])
        add("bronse_vinner", "Bronsevinner", rules["bronse_vinner"] if ok else 0, "riktig" if ok else f'tippet {pred.get("bronse_vinner") or "-"}')
    if fact.get("vm_vinner"):
        ok = _norm(pred.get("vm_vinner")) == _norm(fact["vm_vinner"])
        add("vm_vinner", "Vinner", rules["vm_vinner"] if ok else 0, "riktig" if ok else f'tippet {pred.get("vm_vinner") or "-"}')

    # Bonusfelt avgjøres FØRST ved turneringsslutt. De vises live (med
    # deltakerens tipp), men gir ikke poeng før alt er spilt — slik at en
    # ustabil "ledende" toppscorer/antall mål ikke flytter poeng fram og
    # tilbake mellom de hyppige kjøringene.
    ferdig = bool(fact.get("turnering_ferdig"))

    for fkey, rule_key, label in [
        ("toppscorer", "toppscorer", "Toppscorer"),
        ("assist", "flest_assist", "Flest assist"),
    ]:
        if not fact.get(fkey):
            continue
        tipp = pred.get(fkey) or "-"
        if not ferdig:
            add(fkey, label, 0, f"tippet {tipp} — teller ved slutt")
        else:
            ok = _norm(pred.get(fkey)) == _norm(fact[fkey])
            add(fkey, label, rules[rule_key] if ok else 0, "riktig" if ok else f"tippet {tipp}")

    for fkey, rule_key, label in [
        ("antall_maal", "antall_maal", "Antall mål"),
        ("antall_kort", "antall_kort", "Antall kort"),
    ]:
        if fact.get(fkey) is None:
            continue
        tipp = pred.get(fkey)
        if not ferdig:
            add(fkey, label, 0, f"tippet {tipp} — teller ved slutt")
        else:
            ok = tipp == fact[fkey]
            add(fkey, label, rules[rule_key] if ok else 0, "riktig" if ok else f"tippet {tipp} (fasit {fact[fkey]})")

    return {"total": total, "lines": lines}


# Standard poengregler (fra "Resultater"-arket). Overstyres per turnering
# via tournament.json hvis reglene endres for et framtidig mesterskap.
DEFAULT_RULES = {
    "kamp": 1, "gruppevinner": 3, "r16": 1, "r8": 2, "kvart": 3, "semi": 4,
    "bronse_lag": 1, "bronse_vinner": 3, "finale_lag": 5, "vm_vinner": 10,
    "toppscorer": 10, "flest_assist": 7, "antall_maal": 20, "antall_kort": 20,
}